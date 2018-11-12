import  urllib.parse, uuid, json, requests
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from PIL import Image
from io import BytesIO

search_url = "https://api.cognitive.microsoft.com/bing/v7.0/images/search"
maxTries = 5

class ImageDownloader():

    subscriptionKey = ""
    
    
    def main(self, spreadsheetLocation):   
        if spreadsheetLocation == "":
            spreadsheetLocation = "translatedWords.xlsx"

        self.subscriptionKey = open("imageSubscriptionKey.txt").read()
        print("I'll print each word when I finish downloading a thumbnail for it")

        # Initialize the spreadsheet
        workbook = load_workbook(filename = spreadsheetLocation)
        worksheet = workbook["translations"]

        for rowNum in range(2, worksheet.max_row + 1):
            englishWordCell = "A{}".format(rowNum)
            englishWord = worksheet[englishWordCell].value
            
            found = False
            tryTime = 0
            while not found and tryTime < maxTries:
                try:
                    url = self.getImageUrlForWord(englishWord, tryTime)
                    # Download the image
                    r = requests.get(url)
                    img = Image.open(BytesIO(r.content))
                    found = True
                except OSError:
                    # is okayy
                    print(url + " could not be downloaded.")
                tryTime += 1
                
            # Convert the image into a thumbnail
            size = 256, 256
            img.thumbnail(size, Image.ANTIALIAS)
            
            # Save the thumbnail
            imageFilename = "foodThumbnails/" + englishWord + ".jpg"
            img.save(imageFilename, "JPEG")
            
            print(englishWord)

    def getImageUrlForWord(self, word, retryTime):
        #consider prefacing word with 'edible'
        headers = {"Ocp-Apim-Subscription-Key" : self.subscriptionKey}
        params  = {"q": word}
        response = requests.get(search_url, headers=headers, params=params)
        response.raise_for_status()
        search_results = response.json()
        
        return search_results["value"][retryTime]["contentUrl"]
    
if __name__ == '__main__':
    ImageDownloader().main()