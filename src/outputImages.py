import sys, textwrap
from os.path import join
from openpyxl import load_workbook
from pathlib import Path
from PIL import Image, ImageDraw, ImageFont

thismodule = sys.modules[__name__]

# consts
#Width and height are determined as half A4 paper at 300 dpi
thismodule.imageWidth = 2480
thismodule.imageHeight = 1754
thismodule.horizontalPadding = 50
thismodule.lineSpacingHeight = 100
thismodule.maxFontSize = 200
thismodule.idealFontSize = thismodule.maxFontSize
thismodule.fontLocation = "NotoSansCJKsc-Light.otf"
thismodule.thumbnailSize = (512, 512)

class SpreadsheetWrangler():

    @staticmethod    
    # Creates a dictionary of englishWord: [translatedWord, translatedWord, translatedWord]
    def build_translations_dict (spreadsheetLocation):
        workbook = load_workbook(filename = spreadsheetLocation, read_only=True)
        worksheet = workbook["translations"]

        dict = {}
        
        rows = list(worksheet.rows)
        for row in rows[1:-1]: #start at 1 to skip the header
            firstCell = True
            englishWord = ""
            translations = []
            for cell in row:
                if firstCell:
                    englishWord = cell.value
                    firstCell = False
                else:
                    translations.append(cell.value)

            if englishWord is None: #it's possible to have blank lines
                print("I can't add none to the dictionary!")
                print(translations)
            else:
                dict[englishWord] = translations
        return dict

    @staticmethod
    def get_lists_of_words_per_language(spreadsheetLocation):
        workbook = load_workbook(filename = spreadsheetLocation, read_only=True)
        worksheet = workbook["translations"]

        result = [[], [], [], []]

        rows = list(worksheet.rows)
        for row in rows[1:-1]: #start at 1 to skip the header
            index = 0
            for cell in row:
                if cell.value is not None:
                    result[index].append(cell.value)
                index += 1

        return result

class FinalImageCreater():

    translationsDict = {}
    
    def main(self, spreadsheetLocation):           
        self.translationsDict = SpreadsheetWrangler.build_translations_dict(spreadsheetLocation)

        print("I'll print each word when I finish creating the output image for it")

        for inputWord, translatedWords in self.translationsDict.items():
            if inputWord is None:
                continue

            # Arrange the words into a nice list
            wordsToPrint = translatedWords
            wordsToPrint.insert(0, inputWord)

            # Draw out the white background + the word we translated
            outputImage = Image.new('RGB', (thismodule.imageWidth, thismodule.imageHeight), color = 'white')
            imageDrawer = ImageDraw.Draw(outputImage)            

            # Try to find the thumbnail for this word
            filepath = self.try_get_filepath_for_thumbnail(inputWord)

            # Handle printing the thumbnail if there is one
            if filepath is None:
                foodThumbnailImage = None
                thumbnailWidth, thumbnailHeight = 0,0 
                leftSideOfThumbnail = thismodule.imageWidth
            else: 
                foodThumbnailImage = Image.open(filepath)
                # Ensure the thumbnail is actually thumbnail size
                foodThumbnailImage.thumbnail(thismodule.thumbnailSize, Image.ANTIALIAS)
                # Get the dimensions for the thumbnail
                thumbnailWidth, thumbnailHeight = foodThumbnailImage.size
                leftSideOfThumbnail = thismodule.imageWidth - thumbnailWidth - thismodule.horizontalPadding
                # Paste the thumbnail into the image
                thumbnailOffset = leftSideOfThumbnail, int((thismodule.imageHeight/2) - (thumbnailHeight/2)) # on the right, vertically centered
                outputImage.paste(foodThumbnailImage, thumbnailOffset)

            # The text has a horizontal padding at its left and right, so can't take up all the space before the thumbnail
            textWidth = leftSideOfThumbnail - (thismodule.horizontalPadding  * 2)

            # Calculate the starting point of the text, so that it'll be vertically centered
            thismodule.idealfontsize = FontManipulator.find_font_size_to_fit_height(wordsToPrint, thismodule.imageHeight, thismodule.fontLocation, thismodule.maxFontSize)
            totalHeight = FontManipulator.calculate_total_text_height(wordsToPrint, textWidth, thismodule.fontLocation)
            word_y_pos = int((thismodule.imageHeight - totalHeight) / 2) # e.g. (1754 - 1500)/2 = start at 177

            # Print each word onto the image
            for word in wordsToPrint:
                fontsize = FontManipulator.find_font_size_to_fit_width(word, textWidth, thismodule.fontLocation, thismodule.idealFontSize)
                font = ImageFont.truetype(thismodule.fontLocation, fontsize)
                _, h = FontManipulator.get_font_with_offset(font, word)
                imageDrawer.text((thismodule.horizontalPadding, word_y_pos), word, font=font, fill=(0, 0, 0))
                word_y_pos += h + thismodule.lineSpacingHeight

            # Save off the image
            outputImage.save(join(join("..", "images"), inputWord  + ".png"))

            print(inputWord)
        print("I'm finished creating output images")
		

    # try to find an image for the thumbnail - tries jpg and png. doesn't do anything with casing
    # if it can't find it, it returns empty string
    @staticmethod
    def try_get_filepath_for_thumbnail(word):
        # Try to come up with an appropriate image path
        filepath = join("..", "foodThumbnails")
        filepath = join(filepath, word + ".jpg")	
        if not Path(filepath).is_file():
            filepath = join("..", "foodThumbnails")
            filepath = join(filepath, word + ".png")
            if not Path(filepath).is_file():
                print ("Could not find an appropriate thumbnail for " + word)
                filepath = ""

        return filepath

            
class FontManipulator():
    @staticmethod
    def get_font_with_offset(font, text):
        return map(sum, zip(font.getsize(text), font.getoffset(text)))

    @staticmethod
    def find_font_size_to_fit_height(lines, height, fontLocation, maxSize):
        # If we have to fit 5 lines in 2000 pixels then we have to fit 1 line in 400, so let's just calculate the max size to do that
        height /= len(lines)

        minfontsize = maxSize
        for line in lines:
            linesize = FontManipulator.find_font_size_to_fit(line, height - thismodule.lineSpacingHeight, fontLocation, maxSize, False)
            if linesize < minfontsize:
                minfontsize = linesize

        return minfontsize

    @staticmethod
    def find_font_size_to_fit_width(text, width, fontLocation, maxSize):
        return FontManipulator.find_font_size_to_fit(text, width, fontLocation, maxSize, True)

    @staticmethod
    def find_font_size_to_fit(text, fit_value, fontLocation, maxSize, isWidth):
        # Maybe it's already perfect
        font = ImageFont.truetype(fontLocation, maxSize)
        w, h = FontManipulator.get_font_with_offset(font, text)
        if (isWidth):
            result = w
        else:
            result = h

        if (result < fit_value):
            return maxSize

        # If not, we do a binary search
        size = maxSize
        upperBound = maxSize
        lowerBound = 0
        upperBoundResult = result

        while True:
            font = ImageFont.truetype(fontLocation, size)
            w, h = FontManipulator.get_font_with_offset(font, text)
            if (isWidth):
                result = w
            else:
                result = h
            # If we find the perfect match, return it
            if result == fit_value:
                return size
            # If we blew over the top (say it was 74, and we tried a font size 49 and got 73, and font size 50 and got 75, we'd return the 73)
            elif result > upperBoundResult:
                return lowerBound
            # Business as usual - we have width 74, we got width 50, let's try a bigger font
            elif result > fit_value:
                upperBound = size
                upperBoundResult = result
            # Business as usual - we have width 74, we got width 100, so let's try a smaller font
            else:
                lowerBound = size
            prevsize = size
            size = lowerBound + ((upperBound - lowerBound) / 2)
            size = int(size)
            # Don't want to get caught inf looping
            if size == prevsize:
                return lowerBound

    @staticmethod    
    # Calculates the total text height for a list of words
    def calculate_total_text_height(list, width, fontLocation):
        totalH = 0
        for line in list:
            fontSize = FontManipulator.find_font_size_to_fit_width(line, width, fontLocation, thismodule.idealFontSize)
            font = ImageFont.truetype(fontLocation, fontSize)
            _, h = FontManipulator.get_font_with_offset(font, line)
            totalH += h + thismodule.lineSpacingHeight
        return totalH - thismodule.lineSpacingHeight # we don't need line spacing after the final word
        
if __name__ == '__main__':
    FinalImageCreater().main(join("..", "translatedWords.xlsx"))
