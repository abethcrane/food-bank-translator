import os
from openpyxl import load_workbook, Workbook

class SpreadsheetWrangler():
    @staticmethod
    def write_dict_to_spreadsheet(translationsDict, outputspreadsheet, outputLangNames):
        # Initialize the spreadsheet
        workbook = Workbook(write_only=True)
        worksheet = workbook.create_sheet("translations")
        worksheet.append(["English"] + outputLangNames)

        # Append the list of words  to the spreadsheet
        for inputWord, outputWords in translationsDict.items():
            translatedWords = outputWords
            translatedWords.insert(0, inputWord)
            worksheet.append(translatedWords)

        SpreadsheetWrangler.__safe_save_workbook(workbook, outputspreadsheet)

    @staticmethod
    def get_english_words(spreadsheetLocation):
        # Initialize the spreadsheet
        workbook = SpreadsheetWrangler.__safe_load_workbook(spreadsheetLocation)
        if workbook == None:
            return []
            
        worksheet = workbook["translations"]

        englishWords = []

        for rowNum in range(2, worksheet.max_row + 1):
            englishWordCell = "A{}".format(rowNum)
            englishWord = worksheet[englishWordCell].value

            if englishWord is None or englishWord is "":
                continue

            englishWords.append(englishWord)
        
        return englishWords
        

    @staticmethod    
    # Creates a dictionary of englishWord: [translatedWord, translatedWord, translatedWord]
    def build_translations_dict (spreadsheetLocation):
        workbook = SpreadsheetWrangler.__safe_load_workbook(spreadsheetLocation, read_only=True)
        if workbook == None:
            return {}

        worksheet = workbook["translations"]

        dict = {}
        
        rows = list(worksheet.rows)
        for row in rows[1:]: #start at 1 to skip the header
            firstCell = True
            englishWord = ""
            translations = []
            for cell in row:
                if firstCell:
                    englishWord = cell.value
                    firstCell = False
                else:
                    translations.append(cell.value)

            if englishWord is None: #it's possible to have blank lines
                print("I can't add none to the dictionary!")
                print(translations)
            else:
                dict[englishWord] = translations
        return dict

    @staticmethod
    def get_lists_of_words_per_language(spreadsheetLocation):
        workbook = SpreadsheetWrangler.__safe_load_workbook(spreadsheetLocation, read_only=True)
        if workbook == None:
            return [[], [], [], []]

        worksheet = workbook["translations"]

        result = [[], [], [], []]

        rows = list(worksheet.rows)
        for row in rows[1:-1]: #start at 1 to skip the header
            index = 0
            for cell in row:
                if cell.value is not None:
                    result[index].append(cell.value)
                index += 1

        return result

    @staticmethod
    def __safe_load_workbook(filename, read_only = False):
        if os.path.exists(filename):
            return load_workbook(filename=filename, read_only=read_only)
        else:
            print ("spreadsheet at ", filename, " does not exist")
            return None

    @staticmethod
    # If the directory in the filepath doesn't exist, create it
    def __safe_save_workbook(workbook, filepath):
        parentDir = os.path.dirname(filepath)
        if not os.path.exists(parentDir):
            os.makedirs(parentDir)

        return workbook.save(filepath)

