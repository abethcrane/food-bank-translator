import json, requests, sys, urllib.parse, uuid
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from PIL import Image
from io import BytesIO

thismodule = sys.modules[__name__]

thismodule.search_url = "https://api.cognitive.microsoft.com/bing/v7.0/images/search"
thismodule.maxTries = 5
thismodule.resultsPerQuery = 35
thismodule.thumbnailsize = (512, 512)

class ImageDownloader():

    subscriptionKey = open("../subscriptionKeys/imageSubscriptionKey.txt").read()

    def main(self, spreadsheetLocation):   
        if spreadsheetLocation == "":
            spreadsheetLocation = "../translatedWords.xlsx"

        print("I'll print each word when I finish downloading a thumbnail for it")

        # Initialize the spreadsheet
        workbook = load_workbook(filename = spreadsheetLocation)
        worksheet = workbook["translations"]

        for rowNum in range(2, worksheet.max_row + 1):
            englishWordCell = "A{}".format(rowNum)
            englishWord = worksheet[englishWordCell].value

            if englishWord is None or englishWord is "":
                continue

            (img, _) = self.get_next_image(englishWord, 0)
            if img is None:
                print("I couldn't find an image for " + englishWord)
                continue

            self.thumbify_and_save(englishWord, img)

            print(englishWord)
        
        print("I'm finished downloading thumbnails")

    def get_images_for_words(self, words):
        print ("I'm getting the images for some words - I'll print each one's name as I go")
        for word in words:
            if word is None or word is "":
                continue

            (img, _) = self.get_next_image(word, 0)
            if img is None:
                print("I couldn't find an image for " + word)
                continue

            self.thumbify_and_save(word, img)
            print(word)

    @staticmethod
    def thumbify_and_save(word, img):
        if img is None or word is "":
            print("Cannot save empty image or with empty file names")
            return

        # Convert the image into a thumbnail
        img.thumbnail(thismodule.thumbnailsize, Image.ANTIALIAS)
        
        # Save the thumbnail
        imageFilename = "../foodThumbnails/" + word + ".jpg"
        img.convert('RGB').save(imageFilename, "JPEG")

    def get_image_url_for_word(self, word, retryTime):
        #consider prefacing word with 'edible'
        headers = {"Ocp-Apim-Subscription-Key" : self.subscriptionKey}
        params  = {"q": word}

        try:
            response = requests.get(thismodule.search_url, headers=headers, params=params)
            response.raise_for_status()
            search_results = response.json()
            return search_results["value"][retryTime]["contentUrl"]
        except requests.exceptions.HTTPError:
            print("Could not find results for " + word)
            return ""

    def get_nth_image_for_word(self, word, n):
        url = self.get_image_url_for_word(word, n)
        img = None
        try:
            # Download the image
            r = requests.get(url)
            img = Image.open(BytesIO(r.content))
        except OSError:
            # This is okay, we'll just try another one
            print(url + " could not be downloaded.")

        return img

    def get_next_image(self, word, n):
        if (n < 0):
            n = 0
        img = self.get_nth_image_for_word(word, n)
        retryTime = 0
        while img is None and retryTime < thismodule.maxTries:
            retryTime += 1
            n += 1
            img = self.get_nth_image_for_word(word, n)

        return img, n

    def get_prev_image(self, word, n):
        if (n < 0):
            n = thismodule.resultsPerQuery - 1
        img = self.get_nth_image_for_word(word, n)
        retryTime = 0
        while img is None and retryTime < thismodule.maxTries:
            retryTime += 1
            n -= 1
            img = self.get_nth_image_for_word(word, n)

        return img, n


if __name__ == '__main__':
    ImageDownloader().main("")